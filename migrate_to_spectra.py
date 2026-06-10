#!/usr/bin/env python3
"""
SPECTRA - Script de Migración de Datos Históricos
===================================================
Migra los datos históricos de Costarelli desde Access (exportados a xlsx)
hacia la base de datos Supabase de SPECTRA.

Archivos requeridos en la misma carpeta que este script:
  - CLIENTES.xlsx   (418 clientes)
  - EQUIPOS.xlsx    (4.084 órdenes, NUMPRE 45002-49085)
  - SALIDA.xlsx     (4.025 remitos de salida, 3.499 ítems de trabajo)

Antes de correr:
  1. Ejecutar en Supabase SQL Editor:
     ALTER TABLE ordenes DISABLE TRIGGER trg_set_numero_orden;

Después de correr:
  2. Ejecutar en Supabase SQL Editor:
     ALTER TABLE ordenes ENABLE TRIGGER trg_set_numero_orden;
     -- No hay secuencia: el trigger calcula MAX(numero)+1 directamente.
"""

import os
import sys
from datetime import datetime
from openpyxl import load_workbook
from supabase import create_client, Client

# ─────────────────────────────────────────────
# CONFIGURACIÓN — completar con los valores reales
# ─────────────────────────────────────────────
SUPABASE_URL = os.environ.get('NEXT_PUBLIC_SUPABASE_URL', '')
SUPABASE_SERVICE_KEY = os.environ.get('SUPABASE_SERVICE_ROLE_KEY', '')
EMPRESA_NOMBRE = 'Costarelli Telecomunicaciones'

# Rutas a los xlsx (ajustar si están en otra carpeta)
XLSX_DIR = os.path.dirname(os.path.abspath(__file__))
PATH_CLIENTES = os.path.join(XLSX_DIR, 'CLIENTES.xlsx')
PATH_EQUIPOS  = os.path.join(XLSX_DIR, 'EQUIPOS.xlsx')
PATH_SALIDA   = os.path.join(XLSX_DIR, 'SALIDA.xlsx')

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
EMPTY_VALS = {None, '', '-', 'None', 'none'}

def clean(v):
    """Convierte valores vacíos/guion a None, limpia espacios."""
    if v is None:
        return None
    s = str(v).strip()
    return None if s in EMPTY_VALS else s

def to_bool(v):
    """SI → True, cualquier otra cosa → False."""
    return str(v).strip().upper() == 'SI' if v else False

def parse_fecha_equipos(v):
    """openpyxl ya convierte las fechas de EQUIPOS a datetime."""
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return None

def parse_fecha_salida(v):
    """SALIDA tiene fechas como texto 'DD/MM/YYYY'."""
    s = clean(v)
    if not s:
        return None
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None  # fecha no parseable → NULL (son solo 28 casos)

def normalize_moneda(v):
    """Normaliza los 5 textos distintos de MONEDA a 'ARS' o 'USD'."""
    s = str(v).upper() if v else ''
    if 'PESO' in s:
        return 'ARS'
    if 'DOLAR' in s:
        return 'USD'
    return 'USD'  # default histórico (la mayoría son USD)

def batch_insert(supabase, table, records, size=200):
    """Inserta en batches, devuelve todos los registros insertados."""
    inserted = []
    for i in range(0, len(records), size):
        chunk = records[i:i+size]
        res = supabase.table(table).insert(chunk).execute()
        if res.data:
            inserted.extend(res.data)
    return inserted


# ─────────────────────────────────────────────
# PASO 1 — Obtener empresa_id
# ─────────────────────────────────────────────
def get_empresa_id(supabase):
    # 1) Buscar en 'empresas' (columnas: id, nombre)
    try:
        res = supabase.table('empresas').select('id, nombre').execute()
        for row in res.data or []:
            nombre_col = row.get('nombre') or ''
            if EMPRESA_NOMBRE.lower() in nombre_col.lower():
                eid = row['id']
                print(f'  empresa_id encontrado en "empresas" por nombre: {eid}')
                return eid
        if len(res.data or []) == 1:
            eid = res.data[0]['id']
            print(f'  empresa_id (única en "empresas"): {eid}')
            return eid
    except Exception:
        pass

    # 2) Fallback: buscar en 'config' (columnas: empresa_id, nombre_empresa)
    try:
        res = supabase.table('config').select('empresa_id, nombre_empresa').execute()
        for row in res.data or []:
            nombre_col = row.get('nombre_empresa') or ''
            if EMPRESA_NOMBRE.lower() in nombre_col.lower():
                eid = row['empresa_id']
                print(f'  empresa_id encontrado en "config" por nombre: {eid}')
                return eid
        if len(res.data or []) == 1:
            eid = res.data[0]['empresa_id']
            print(f'  empresa_id (única en "config"): {eid}')
            return eid
    except Exception:
        pass

    raise Exception(
        f'No se pudo obtener empresa_id. Revisá el nombre exacto "{EMPRESA_NOMBRE}" '
        'en la tabla config/empresas de Supabase.'
    )


# ─────────────────────────────────────────────
# PASO 2 — Migrar CLIENTES
# ─────────────────────────────────────────────
def migrate_clientes(supabase, empresa_id):
    """
    Inserta 418 clientes. Devuelve dict {access_id (str): supabase_uuid}.
    Estrategia: inserta de a 1 para capturar UUID retornado directamente.
    """
    wb = load_workbook(PATH_CLIENTES, read_only=True)
    rows = [r for r in wb.active.iter_rows(min_row=2, values_only=True) if r[0]]
    
    access_id_to_uuid = {}
    skipped = 0

    for r in rows:
        access_id = str(r[0])
        nombre = clean(r[1])
        if not nombre:  # 7 clientes sin nombre → skip
            skipped += 1
            continue
        
        record = {
            'empresa_id':    empresa_id,
            'razon_social':  nombre,
            'direccion':     clean(r[2]),
            'codigo_postal': clean(r[3]),
            'provincia':     clean(r[4]),
            'localidad':     clean(r[5]),
            'contacto':      clean(r[6]),
            'telefono1':     clean(r[7]),
            # r[8] (mail) y r[9] (cuit) no tienen columna en clientes → se descartan
        }
        
        try:
            res = supabase.table('clientes').insert(record).execute()
            if res.data:
                access_id_to_uuid[access_id] = res.data[0]['id']
        except Exception as e:
            # Si ya existe (unique constraint), buscar por nombre
            try:
                existing = (supabase.table('clientes')
                            .select('id')
                            .eq('empresa_id', empresa_id)
                            .eq('razon_social', nombre)
                            .limit(1)
                            .execute())
                if existing.data:
                    access_id_to_uuid[access_id] = existing.data[0]['id']
            except Exception:
                pass

    print(f'  Clientes insertados: {len(access_id_to_uuid)} | Skipped (sin nombre): {skipped}')
    return access_id_to_uuid


# ─────────────────────────────────────────────
# PASO 3 — Mapear/crear técnicos
# ─────────────────────────────────────────────
def sync_tecnicos(supabase, empresa_id):
    """Levanta técnicos existentes, crea los nuevos del histórico."""
    wb = load_workbook(PATH_SALIDA, read_only=True)
    names_raw = set()
    for r in wb.active.iter_rows(min_row=2, values_only=True):
        n = clean(r[44])  # columna TENCNICO
        if n:
            names_raw.add(n.strip())

    # Técnicos existentes en SPECTRA
    existing = supabase.table('tecnicos').select('id,nombre').eq('empresa_id', empresa_id).execute().data or []
    tec_map = {t['nombre'].upper(): t['id'] for t in existing}

    created = 0
    for name in sorted(names_raw):
        key = name.upper()
        if key not in tec_map:
            try:
                res = supabase.table('tecnicos').insert({'empresa_id': empresa_id, 'nombre': name}).execute()
                if res.data:
                    tec_map[key] = res.data[0]['id']
                    created += 1
            except Exception:
                pass

    print(f'  Técnicos: {len(existing)} existentes + {created} nuevos = {len(tec_map)} total')
    return tec_map


# ─────────────────────────────────────────────
# PASO 4 — Migrar EQUIPOS (tabla ordenes)
# ─────────────────────────────────────────────
def migrate_equipos(supabase, empresa_id, access_id_to_uuid, tec_map):
    """
    IMPORTANTE: Antes de correr este paso, deshabilitar el trigger
    que auto-asigna el número correlativo en la tabla ordenes.
    Ver instrucciones al inicio del script.

    Los accesorios (microfono, fuente, etc.) viven en la tabla separada
    accesorios_orden (1:1 con ordenes), por eso se insertan en dos pasos.
    """
    wb = load_workbook(PATH_EQUIPOS, read_only=True)
    rows = [r for r in wb.active.iter_rows(min_row=2, values_only=True) if r[0]]

    ordenes_batch       = []
    accesorios_by_numpre = {}   # numpre → dict con los 7 campos de accesorios
    sin_cliente = 0

    for r in rows:
        numpre        = r[0]
        access_cli_id = clean(r[16])  # columna Id (FK a CLIENTES)
        cliente_uuid  = access_id_to_uuid.get(access_cli_id) if access_cli_id else None
        if not cliente_uuid:
            sin_cliente += 1
            continue   # cliente_id es NOT NULL — saltear esta orden

        entrego_val = clean(r[4])
        # En EQUIPOS el campo ENTREGO está casi siempre vacío (98%).
        # El estado real "entregado" se actualiza en PASO 5 desde SALIDA.
        estado = 'entregado' if (entrego_val and entrego_val != '-') else 'ingresado'

        ordenes_batch.append({
            'empresa_id':    empresa_id,
            'numero':        numpre,            # ← requiere trigger deshabilitado
            'cliente_id':    cliente_uuid,
            'marca':         clean(r[1]),
            'modelo':        clean(r[2]),
            'numero_serie':  clean(r[3]),
            'estacion':      clean(r[6]),
            'deficiencia':   clean(r[5]),       # columna DETALLE = campo deficiencia
            'observaciones': clean(r[7]),
            'fecha_ingreso': parse_fecha_equipos(r[15]),
            'estado':        estado,
        })

        # Los accesorios van a la tabla accesorios_orden (1:1 con ordenes)
        accesorios_by_numpre[numpre] = {
            'microfono': to_bool(r[8]),
            'fuente':    to_bool(r[9]),
            'cable':     to_bool(r[10]),
            'pack':      to_bool(r[11]),
            'antena':    to_bool(r[12]),
            'cargador':  to_bool(r[13]),
            'crem':      to_bool(r[14]),
        }

    inserted = batch_insert(supabase, 'ordenes', ordenes_batch)

    # Construir mapa NUMPRE → UUID — paginar de a 1000 para superar el límite de PostgREST
    all_ord = []
    page = 0
    page_size = 1000
    while True:
        chunk = (supabase.table('ordenes')
                 .select('id,numero')
                 .eq('empresa_id', empresa_id)
                 .range(page * page_size, (page + 1) * page_size - 1)
                 .execute().data or [])
        all_ord.extend(chunk)
        if len(chunk) < page_size:
            break
        page += 1
    numpre_to_uuid = {str(o['numero']): o['id'] for o in all_ord}

    # Insertar accesorios_orden usando los UUIDs recién obtenidos
    acc_batch = []
    for numpre, acc in accesorios_by_numpre.items():
        orden_uuid = numpre_to_uuid.get(str(numpre))
        if orden_uuid:
            acc_batch.append({'orden_id': orden_uuid, **acc})
    batch_insert(supabase, 'accesorios_orden', acc_batch)

    print(f'  Órdenes insertadas: {len(inserted)} | Sin cliente: {sin_cliente} | Mapeadas: {len(numpre_to_uuid)}')
    print(f'  Accesorios insertados: {len(acc_batch)}')
    return numpre_to_uuid


# ─────────────────────────────────────────────
# PASO 5 — Migrar SALIDA (ítems + actualizar ordenes)
# ─────────────────────────────────────────────
def migrate_salida(supabase, empresa_id, numpre_to_uuid, tec_map):
    wb = load_workbook(PATH_SALIDA, read_only=True)
    rows = [r for r in wb.active.iter_rows(min_row=2, values_only=True) if r[0]]

    items_batch    = []
    ordenes_update = {}   # orden_uuid → dict con campos a actualizar
    sin_match      = 0

    for r in rows:
        npresentrada = clean(r[1])
        orden_uuid   = numpre_to_uuid.get(str(npresentrada)) if npresentrada else None

        if not orden_uuid:
            sin_match += 1
            continue

        # ── Datos de salida para actualizar la orden ──────────────────────
        entrego      = clean(r[43])
        tecnico_name = clean(r[44])
        tecnico_id   = tec_map.get(tecnico_name.upper()) if tecnico_name else None
        fecha_egreso = parse_fecha_salida(r[55])
        moneda       = normalize_moneda(r[46])
        iva_text     = str(r[45]) if r[45] else ''
        iva          = 'INCLUYE' in iva_text.upper()

        if orden_uuid not in ordenes_update:
            update_payload = {}
            if entrego:      update_payload['entrego']     = entrego
            if tecnico_id:   update_payload['tecnico']     = tecnico_name   # text, no FK
            if fecha_egreso: update_payload['fecha_salida'] = fecha_egreso
            update_payload['moneda']     = moneda
            update_payload['aplica_iva'] = iva
            if entrego:      update_payload['estado'] = 'entregado'
            ordenes_update[orden_uuid] = update_payload

        # ── Ítems de trabajo (hasta 7 renglones por remito) ───────────────
        for i in range(7):
            qty   = r[22 + i * 3]
            det   = r[23 + i * 3]
            price = r[24 + i * 3]

            qty_s   = str(qty).strip()   if qty   else ''
            det_s   = str(det).strip()   if det   else ''
            price_s = str(price).strip() if price else '0'

            if qty_s in ('', '0', '-') or det_s in ('', '-'):
                continue

            try:
                items_batch.append({
                    'orden_id': orden_uuid,
                    'cantidad': int(float(qty_s)),
                    'detalle':  det_s,
                    'precio':   float(price_s.replace(',', '.')) if price_s not in ('0', '-', '') else 0.0,
                })
            except (ValueError, TypeError):
                pass

    # Insertar ítems
    items_inserted = batch_insert(supabase, 'items_trabajo', items_batch)

    # Actualizar órdenes (update individual — ~3.500 registros, rápido)
    updates_ok = 0
    for uuid, payload in ordenes_update.items():
        if payload:
            try:
                supabase.table('ordenes').update(payload).eq('id', uuid).execute()
                updates_ok += 1
            except Exception:
                pass

    print(f'  Ítems de trabajo insertados: {len(items_inserted)}')
    print(f'  Órdenes actualizadas con datos de salida: {updates_ok}')
    print(f'  Remitos sin match de orden: {sin_match}')


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    if not SUPABASE_URL or not SUPABASE_SERVICE_KEY:
        print('\nERROR: Variables de entorno no configuradas.')
        print('   Necesitás: NEXT_PUBLIC_SUPABASE_URL y SUPABASE_SERVICE_ROLE_KEY')
        print('   (las encontrás en el archivo .env.local del proyecto)')
        sys.exit(1)

    print(f'\n{"=" * 55}')
    print('  SPECTRA — Migración de datos históricos de Costarelli')
    print(f'{"=" * 55}\n')

    supabase: Client = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

    print('[1/5] Obteniendo empresa_id...')
    empresa_id = get_empresa_id(supabase)

    print('[2/5] Migrando clientes...')
    access_id_to_uuid = migrate_clientes(supabase, empresa_id)

    print('[3/5] Sincronizando técnicos...')
    tec_map = sync_tecnicos(supabase, empresa_id)

    print('[4/5] Migrando ordenes (EQUIPOS)...')
    print('  (trigger trg_set_numero_orden debe estar deshabilitado en Supabase)')
    numpre_to_uuid = migrate_equipos(supabase, empresa_id, access_id_to_uuid, tec_map)

    print('[5/5] Migrando salidas e ítems de trabajo...')
    migrate_salida(supabase, empresa_id, numpre_to_uuid, tec_map)

    print(f'\n{"=" * 55}')
    print('  MIGRACION COMPLETADA')
    print(f'{"=" * 55}')
    print("""
Pasos finales en Supabase SQL Editor:

  -- 1. Rehabilitar el trigger de numeración:
  ALTER TABLE ordenes ENABLE TRIGGER trg_set_numero_orden;
  -- El trigger usa MAX(numero)+1, no hay secuencia que actualizar.
  -- La próxima orden que se ingrese tomará MAX(numero)+1 automáticamente.

  -- 2. Verificar:
  SELECT COUNT(*) FROM clientes        WHERE empresa_id = '<empresa_id>';
  SELECT COUNT(*) FROM ordenes         WHERE empresa_id = '<empresa_id>';
  SELECT COUNT(*) FROM accesorios_orden WHERE orden_id IN (SELECT id FROM ordenes WHERE empresa_id = '<empresa_id>');
  SELECT COUNT(*) FROM items_trabajo   WHERE orden_id  IN (SELECT id FROM ordenes WHERE empresa_id = '<empresa_id>');
""")


if __name__ == '__main__':
    main()
