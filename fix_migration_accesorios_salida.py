#!/usr/bin/env python3
"""
SPECTRA - Fix de migración: accesorios_orden e items_trabajo
=============================================================
Corre DESPUÉS de migrate_to_spectra.py cuando la query de ordenes quedó
truncada a 1000 filas (límite de PostgREST).

Este script:
  1. Lee TODOS los ordenes de la BD (paginado de a 1000)
  2. Borra accesorios_orden e items_trabajo incompletos de esta empresa
  3. Re-inserta accesorios_orden desde EQUIPOS.xlsx
  4. Re-inserta items_trabajo y actualiza ordenes desde SALIDA.xlsx

Pre-requisito: el trigger trg_set_numero_orden debe seguir deshabilitado
(ya fue deshabilitado antes de migrate_to_spectra.py).
"""

import os
import sys
from datetime import datetime
from openpyxl import load_workbook
from supabase import create_client, Client

SUPABASE_URL      = os.environ.get('NEXT_PUBLIC_SUPABASE_URL', '')
SUPABASE_SERVICE_KEY = os.environ.get('SUPABASE_SERVICE_ROLE_KEY', '')
EMPRESA_NOMBRE    = 'Costarelli Telecomunicaciones'

XLSX_DIR      = os.path.dirname(os.path.abspath(__file__))
PATH_EQUIPOS  = os.path.join(XLSX_DIR, 'EQUIPOS.xlsx')
PATH_SALIDA   = os.path.join(XLSX_DIR, 'SALIDA.xlsx')

EMPTY_VALS = {None, '', '-', 'None', 'none'}

def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return None if s in EMPTY_VALS else s

def to_bool(v):
    return str(v).strip().upper() == 'SI' if v else False

def parse_fecha_salida(v):
    s = clean(v)
    if not s:
        return None
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None

def normalize_moneda(v):
    s = str(v).upper() if v else ''
    if 'PESO' in s:
        return 'ARS'
    if 'DOLAR' in s:
        return 'USD'
    return 'USD'

def batch_insert(supabase, table, records, size=200):
    inserted = 0
    for i in range(0, len(records), size):
        chunk = records[i:i+size]
        res = supabase.table(table).insert(chunk).execute()
        inserted += len(res.data or [])
    return inserted

def get_empresa_id(supabase):
    res = supabase.table('empresas').select('id, nombre').execute()
    for row in res.data or []:
        if EMPRESA_NOMBRE.lower() in (row.get('nombre') or '').lower():
            return row['id']
    if len(res.data or []) == 1:
        return res.data[0]['id']
    raise Exception(f'empresa_id no encontrado para "{EMPRESA_NOMBRE}"')

def fetch_all_ordenes(supabase, empresa_id):
    """Trae todos los ordenes paginando de a 1000 (límite de PostgREST)."""
    all_ord = []
    page_size = 1000
    page = 0
    while True:
        chunk = (supabase.table('ordenes')
                 .select('id,numero')
                 .eq('empresa_id', empresa_id)
                 .range(page * page_size, (page + 1) * page_size - 1)
                 .execute().data or [])
        all_ord.extend(chunk)
        print(f'  Página {page + 1}: {len(chunk)} ordenes (total acumulado: {len(all_ord)})')
        if len(chunk) < page_size:
            break
        page += 1
    return {str(o['numero']): o['id'] for o in all_ord}

def delete_existing_data(supabase, empresa_id, numpre_to_uuid):
    """Borra accesorios_orden e items_trabajo existentes para esta empresa."""
    orden_ids = list(numpre_to_uuid.values())
    # Borrar en lotes de 100 IDs para no superar límite de URL
    deleted_acc = 0
    deleted_items = 0
    for i in range(0, len(orden_ids), 100):
        chunk_ids = orden_ids[i:i+100]
        res = supabase.table('accesorios_orden').delete().in_('orden_id', chunk_ids).execute()
        deleted_acc += len(res.data or [])
        res = supabase.table('items_trabajo').delete().in_('orden_id', chunk_ids).execute()
        deleted_items += len(res.data or [])
    print(f'  Borrados: {deleted_acc} accesorios_orden, {deleted_items} items_trabajo')

def reinsert_accesorios(supabase, numpre_to_uuid):
    """Re-inserta accesorios_orden desde EQUIPOS.xlsx."""
    wb = load_workbook(PATH_EQUIPOS, read_only=True)
    rows = [r for r in wb.active.iter_rows(min_row=2, values_only=True) if r[0]]

    acc_batch = []
    sin_orden = 0
    for r in rows:
        numpre = str(r[0])
        orden_uuid = numpre_to_uuid.get(numpre)
        if not orden_uuid:
            sin_orden += 1
            continue
        acc_batch.append({
            'orden_id': orden_uuid,
            'microfono': to_bool(r[8]),
            'fuente':    to_bool(r[9]),
            'cable':     to_bool(r[10]),
            'pack':      to_bool(r[11]),
            'antena':    to_bool(r[12]),
            'cargador':  to_bool(r[13]),
            'crem':      to_bool(r[14]),
        })

    inserted = batch_insert(supabase, 'accesorios_orden', acc_batch)
    print(f'  Accesorios insertados: {inserted} | Sin orden en BD: {sin_orden}')

def rerun_salida(supabase, empresa_id, numpre_to_uuid):
    """Re-inserta items_trabajo y actualiza ordenes desde SALIDA.xlsx."""
    wb = load_workbook(PATH_SALIDA, read_only=True)
    rows = [r for r in wb.active.iter_rows(min_row=2, values_only=True) if r[0]]

    items_batch    = []
    ordenes_update = {}
    sin_match      = 0

    for r in rows:
        npresentrada = clean(r[1])
        orden_uuid   = numpre_to_uuid.get(str(npresentrada)) if npresentrada else None
        if not orden_uuid:
            sin_match += 1
            continue

        entrego      = clean(r[43])
        tecnico_name = clean(r[44])
        fecha_egreso = parse_fecha_salida(r[55])
        moneda       = normalize_moneda(r[46])
        iva_text     = str(r[45]) if r[45] else ''
        iva          = 'INCLUYE' in iva_text.upper()

        if orden_uuid not in ordenes_update:
            payload = {'moneda': moneda, 'aplica_iva': iva}
            if entrego:      payload['entrego']     = entrego
            if tecnico_name: payload['tecnico']     = tecnico_name
            if fecha_egreso: payload['fecha_salida'] = fecha_egreso
            if entrego:      payload['estado']      = 'entregado'
            ordenes_update[orden_uuid] = payload

        for i in range(7):
            qty   = r[22 + i * 3]
            det   = r[23 + i * 3]
            price = r[24 + i * 3]
            qty_s   = str(qty).strip()   if qty   else ''
            det_s   = str(det).strip()   if det   else ''
            price_s = str(price).strip() if price else '0'
            if qty_s in ('', '0', '-') or det_s in ('', '-'):
                continue
            try:
                items_batch.append({
                    'orden_id': orden_uuid,
                    'cantidad': int(float(qty_s)),
                    'detalle':  det_s,
                    'precio':   float(price_s.replace(',', '.')) if price_s not in ('0', '-', '') else 0.0,
                })
            except (ValueError, TypeError):
                pass

    items_inserted = batch_insert(supabase, 'items_trabajo', items_batch)

    updates_ok = 0
    for uuid, payload in ordenes_update.items():
        if payload:
            try:
                supabase.table('ordenes').update(payload).eq('id', uuid).execute()
                updates_ok += 1
            except Exception:
                pass

    print(f'  Items de trabajo insertados: {items_inserted}')
    print(f'  Ordenes actualizadas con salida: {updates_ok}')
    print(f'  Remitos sin match de orden: {sin_match}')


def main():
    if not SUPABASE_URL or not SUPABASE_SERVICE_KEY:
        print('ERROR: Variables de entorno no configuradas.')
        sys.exit(1)

    print(f'\n{"=" * 55}')
    print('  SPECTRA — Fix accesorios_orden + items_trabajo')
    print(f'{"=" * 55}\n')

    supabase: Client = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

    print('[1/4] Obteniendo empresa_id...')
    empresa_id = get_empresa_id(supabase)
    print(f'  empresa_id: {empresa_id}')

    print('[2/4] Cargando TODOS los ordenes de la BD (paginado)...')
    numpre_to_uuid = fetch_all_ordenes(supabase, empresa_id)
    print(f'  Total ordenes mapeados: {len(numpre_to_uuid)}')

    print('[3/4] Limpiando accesorios e items anteriores...')
    delete_existing_data(supabase, empresa_id, numpre_to_uuid)

    print('[4a/4] Re-insertando accesorios_orden...')
    reinsert_accesorios(supabase, numpre_to_uuid)

    print('[4b/4] Re-procesando salidas e items de trabajo...')
    rerun_salida(supabase, empresa_id, numpre_to_uuid)

    print(f'\n{"=" * 55}')
    print('  FIX COMPLETADO')
    print(f'{"=" * 55}')
    print(f"""
Pasos finales en Supabase SQL Editor:

  ALTER TABLE ordenes ENABLE TRIGGER trg_set_numero_orden;

  -- Verificar totales:
  SELECT COUNT(*) FROM clientes         WHERE empresa_id = '{empresa_id}';
  SELECT COUNT(*) FROM ordenes          WHERE empresa_id = '{empresa_id}';
  SELECT COUNT(*) FROM accesorios_orden WHERE orden_id IN (SELECT id FROM ordenes WHERE empresa_id = '{empresa_id}');
  SELECT COUNT(*) FROM items_trabajo    WHERE orden_id  IN (SELECT id FROM ordenes WHERE empresa_id = '{empresa_id}');
""")


if __name__ == '__main__':
    main()
